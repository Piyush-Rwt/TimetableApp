import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
import os
import sys

if getattr(sys, 'frozen', False):
    BASE_DIR = os.path.dirname(sys.executable)
else:
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

DEFAULT_EXPORT_PATH = os.path.join(BASE_DIR, "saved_tt", "University_Timetable.xlsx")

def export_full_timetable(result, sections, subjects, teachers, rooms, filepath=DEFAULT_EXPORT_PATH):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    sub_map = {s['id']: s for s in subjects}
    t_map = {t['id']: t['name'] for t in teachers}
    r_map = {r['id']: r['name'] for r in rooms}
    
    days = result['days']
    slots_times = result['slots_times']
    num_slots = len(slots_times)
    grid = result['grid']
    
    # Styles
    header_fill = PatternFill(start_color="4F6EF7", end_color="4F6EF7", fill_type="solid")
    break_fill = PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid")
    lab_fill = PatternFill(start_color="2ECC71", end_color="2ECC71", fill_type="solid")
    theory_fill = PatternFill(start_color="A9CCE3", end_color="A9CCE3", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    def create_sheet(title, data_provider):
        ws = wb.create_sheet(title=title[:31]) # Excel sheet name limit
        
        # Headers
        ws.cell(row=1, column=1).value = "Time / Day"
        ws.cell(row=1, column=1).fill = header_fill
        ws.cell(row=1, column=1).font = Font(color="FFFFFF", bold=True)
        ws.cell(row=1, column=1).border = thin_border
        
        for col_idx, day in enumerate(days, 2):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = day
            cell.fill = header_fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
            
        # Rows
        for slot_idx in range(num_slots):
            row_idx = slot_idx + 2
            t_range = f"{slots_times[slot_idx][0]} - {slots_times[slot_idx][1]}"
            ws.cell(row=row_idx, column=1).value = t_range
            ws.cell(row=row_idx, column=1).border = thin_border
            
            for col_idx, day in enumerate(days, 2):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                entry = data_provider(day, slot_idx)
                if entry == "BREAK":
                    cell.value = "BREAK"
                    cell.fill = break_fill
                elif entry:
                    # entry is {subject_id, teacher_id, room_id, is_lab}
                    sub = sub_map.get(entry['subject_id'], {})
                    sub_name = sub.get('name', 'Unknown')
                    t_name = t_map.get(entry['teacher_id'], 'Unknown')
                    r_name = r_map.get(entry['room_id'], 'Unknown')
                    
                    cell.value = f"{sub_name}\n{t_name}\n[{r_name}]"
                    cell.fill = lab_fill if entry.get('is_lab') else theory_fill
                else:
                    cell.value = ""

    # 1. Section Sheets
    for sec in sections:
        def section_data(day, slot):
            return grid.get(sec['id'], {}).get(day, {}).get(slot)
        create_sheet(f"Sec {sec['name']}", section_data)

    # 2. Teacher Sheets (optional, but requested)
    for t_id, t_name in t_map.items():
        def teacher_data(day, slot):
            # Find if teacher is busy in any section grid at this slot
            for sec_id, days_grid in grid.items():
                entry = days_grid.get(day, {}).get(slot)
                if entry and entry != "BREAK" and entry.get('teacher_id') == t_id:
                    # Add section info to the display
                    sec_name = next(s['name'] for s in sections if s['id'] == sec_id)
                    return {**entry, 'section_name': sec_name}
            # Check if it was a break slot globally
            first_sec_id = sections[0]['id']
            if grid[first_sec_id][day][slot] == "BREAK":
                return "BREAK"
            return None
        # create_sheet(f"Tea {t_name}", teacher_data) # Only if teacher has classes?

    # 3. Room Sheets
    for r_id, r_name in r_map.items():
        def room_data(day, slot):
            for sec_id, days_grid in grid.items():
                entry = days_grid.get(day, {}).get(slot)
                if entry and entry != "BREAK" and entry.get('room_id') == r_id:
                    sec_name = next(s['name'] for s in sections if s['id'] == sec_id)
                    return {**entry, 'section_name': sec_name}
            first_sec_id = sections[0]['id']
            if grid[first_sec_id][day][slot] == "BREAK":
                return "BREAK"
            return None
        # create_sheet(f"Room {r_name}", room_data)

    wb.save(filepath)
    return filepath
