from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, 
    QCheckBox, QSpinBox, QPushButton, QTableWidget, QTableWidgetItem, QHeaderView, QFileDialog, QMessageBox
)
from PySide6.QtGui import QColor
import openpyxl

class SubjectSetupScreen(QWidget):
    def __init__(self, wizard_cb):
        super().__init__()
        self.wizard_cb = wizard_cb
        layout = QVBoxLayout(self)

        title = QLabel("Step 4: Subject Setup")
        title.setObjectName("TitleLabel")
        layout.addWidget(title)

        # Excel Buttons for Subjects
        sub_btn_layout = QHBoxLayout()
        btn_dl_sub = QPushButton("Download Subject Template")
        btn_dl_sub.clicked.connect(self.download_subject_template)
        btn_up_sub = QPushButton("Upload Subject Excel")
        btn_up_sub.clicked.connect(self.upload_subject_excel)
        sub_btn_layout.addWidget(btn_dl_sub)
        sub_btn_layout.addWidget(btn_up_sub)
        sub_btn_layout.addStretch()
        layout.addLayout(sub_btn_layout)

        # Part A
        layout.addWidget(QLabel("Subject Definition"))
        self.tbl_subjects = QTableWidget(0, 7)
        self.tbl_subjects.setHorizontalHeaderLabels([
            "Code", "Name", "Hours/Week", "Is Lab", "Lab Slots", "Special Room", "Assigned Groups"
        ])
        self.tbl_subjects.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        layout.addWidget(self.tbl_subjects)

        btn_add_sub = QPushButton("Add Subject Row")
        btn_add_sub.clicked.connect(self.add_sub_row)
        layout.addWidget(btn_add_sub)

        # Part D (Room Setup)
        layout.addSpacing(20)
        room_header = QHBoxLayout()
        room_header.addWidget(QLabel("Room Setup"))
        btn_up_room = QPushButton("Upload Rooms Excel")
        btn_up_room.clicked.connect(self.upload_rooms_excel)
        room_header.addWidget(btn_up_room)
        room_header.addStretch()
        layout.addLayout(room_header)

        self.tbl_rooms = QTableWidget(0, 4)
        self.tbl_rooms.setHorizontalHeaderLabels(["Room ID", "Room Name", "Type", "Capacity"])
        self.tbl_rooms.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        layout.addWidget(self.tbl_rooms)

        btn_add_room = QPushButton("Add Room Row")
        btn_add_room.clicked.connect(lambda: self.tbl_rooms.insertRow(self.tbl_rooms.rowCount()))
        layout.addWidget(btn_add_room)

    def download_subject_template(self):
        path, _ = QFileDialog.getSaveFileName(self, "Save Template", "saved_tt/Subject_Template.xlsx", "Excel Files (*.xlsx)")
        if path:
            wb = openpyxl.Workbook()
            ws = wb.active
            headers = ["Code", "SubjectName", "HoursPerWeek", "IsLab", "LabDurationSlots", "RequiresSpecialRoom", "RoomType", "AssignedGroups"]
            ws.append(headers)
            ws.append(["AI201", "Machine Learning", 4, "No", 1, "No", "Classroom", "AI/ML"])
            ws.append(["AI203", "Deep Learning Lab", 2, "Yes", 3, "Yes", "AI Lab", "AI/ML"])
            wb.save(path)
            QMessageBox.information(self, "Success", "Template saved successfully.")

    def upload_subject_excel(self):
        path, _ = QFileDialog.getOpenFileName(self, "Open Subject Excel", "saved_tt", "Excel Files (*.xlsx)")
        if path:
            wb = openpyxl.load_workbook(path)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            count = 0
            if len(rows) > 1:
                for row in rows[1:]:
                    if not row[0]: continue
                    r = self.tbl_subjects.rowCount()
                    self.tbl_subjects.insertRow(r)
                    
                    self.tbl_subjects.setItem(r, 0, QTableWidgetItem(str(row[0])))
                    self.tbl_subjects.setItem(r, 1, QTableWidgetItem(str(row[1])))
                    self.tbl_subjects.setItem(r, 2, QTableWidgetItem(str(row[2])))
                    
                    chk_lab = QCheckBox()
                    chk_lab.setChecked(str(row[3]).lower() in ["yes", "true", "1"])
                    self.tbl_subjects.setCellWidget(r, 3, chk_lab)
                    
                    spn_lab = QSpinBox()
                    spn_lab.setValue(int(row[4] or 1))
                    self.tbl_subjects.setCellWidget(r, 4, spn_lab)
                    
                    chk_spec = QCheckBox()
                    chk_spec.setChecked(str(row[5]).lower() in ["yes", "true", "1"])
                    self.tbl_subjects.setCellWidget(r, 5, chk_spec)
                    
                    self.tbl_subjects.setItem(r, 6, QTableWidgetItem(str(row[7])))
                    count += 1
            QMessageBox.information(self, "Success", f"{count} subjects loaded from Excel.")

    def upload_rooms_excel(self):
        path, _ = QFileDialog.getOpenFileName(self, "Open Rooms Excel", "saved_tt", "Excel Files (*.xlsx)")
        if path:
            wb = openpyxl.load_workbook(path)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            count = 0
            if len(rows) > 1:
                for row in rows[1:]:
                    if not row[0]: continue
                    r = self.tbl_rooms.rowCount()
                    self.tbl_rooms.insertRow(r)
                    for c_idx, val in enumerate(row[:4]):
                        self.tbl_rooms.setItem(r, c_idx, QTableWidgetItem(str(val) if val is not None else ""))
                    count += 1
            QMessageBox.information(self, "Success", f"{count} rooms loaded from Excel.")

    def add_sub_row(self):
        r = self.tbl_subjects.rowCount()
        self.tbl_subjects.insertRow(r)
        
        chk_lab = QCheckBox()
        self.tbl_subjects.setCellWidget(r, 3, chk_lab)
        
        spn_lab = QSpinBox()
        spn_lab.setRange(1, 3)
        spn_lab.setValue(1)
        self.tbl_subjects.setCellWidget(r, 4, spn_lab)
        
        chk_spec = QCheckBox()
        self.tbl_subjects.setCellWidget(r, 5, chk_spec)

    def save_data(self):
        pass
