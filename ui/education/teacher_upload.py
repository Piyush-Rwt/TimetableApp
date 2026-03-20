import json
from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton, 
    QTableWidget, QTableWidgetItem, QFileDialog, QMessageBox, QHeaderView
)
from db.queries import insert_teacher, delete_all_teachers
import openpyxl

class TeacherUploadScreen(QWidget):
    def __init__(self, wizard_cb):
        super().__init__()
        self.wizard_cb = wizard_cb
        layout = QVBoxLayout(self)

        title = QLabel("Step 1: Teacher Upload")
        title.setObjectName("TitleLabel")
        layout.addWidget(title)

        btn_layout = QHBoxLayout()
        btn_dl = QPushButton("Download Teacher Template")
        btn_dl.clicked.connect(self.download_template)
        btn_up = QPushButton("Upload Teacher Excel")
        btn_up.clicked.connect(self.upload_excel)
        btn_layout.addWidget(btn_dl)
        btn_layout.addWidget(btn_up)
        btn_layout.addStretch()
        layout.addLayout(btn_layout)

        self.table = QTableWidget(0, 10)
        self.table.setHorizontalHeaderLabels([
            "TeacherID", "Name", "Department", "Subjects", 
            "MaxTeachHoursPerDay", "MaxStayHoursPerDay", 
            "AvailableFrom", "AvailableTill", "UnavailableDays", "Notes"
        ])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        layout.addWidget(self.table)

        btn_save = QPushButton("Save Changes")
        btn_save.setObjectName("PrimaryButton")
        btn_save.clicked.connect(self.save_data)
        layout.addWidget(btn_save)

    def download_template(self):
        path, _ = QFileDialog.getSaveFileName(self, "Save Template", "saved_tt/Teacher_Template.xlsx", "Excel Files (*.xlsx)")
        if path:
            wb = openpyxl.Workbook()
            ws = wb.active
            headers = ["TeacherID", "Name", "Department", "Subjects", "MaxTeachHoursPerDay", "MaxStayHoursPerDay", "AvailableFrom", "AvailableTill", "UnavailableDays", "Notes"]
            ws.append(headers)
            ws.append(["T001", "Dr. Mehta", "CSE", "AI,ML,Python", 4, 6, "12:00", "18:00", "Saturday", "HOD, only after noon"])
            wb.save(path)
            QMessageBox.information(self, "Success", "Template saved successfully.")

    def upload_excel(self):
        path, _ = QFileDialog.getOpenFileName(self, "Open Teacher Excel", "saved_tt", "Excel Files (*.xlsx)")
        if path:
            wb = openpyxl.load_workbook(path)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) > 1:
                self.table.setRowCount(0)
                for r_idx, row in enumerate(rows[1:]):
                    if not row[1]: continue # Skip if no name
                    self.table.insertRow(r_idx)
                    for c_idx, val in enumerate(row):
                        self.table.setItem(r_idx, c_idx, QTableWidgetItem(str(val) if val is not None else ""))

    def save_data(self):
        delete_all_teachers()
        
        for r in range(self.table.rowCount()):
            data = {
                "name": self.table.item(r, 1).text() if self.table.item(r, 1) else "",
                "department": self.table.item(r, 2).text() if self.table.item(r, 2) else "",
                "subjects": self.table.item(r, 3).text() if self.table.item(r, 3) else "",
                "max_teach_hrs": int(self.table.item(r, 4).text() or 5),
                "max_stay_hrs": int(self.table.item(r, 5).text() or 8),
                "available_from": self.table.item(r, 6).text() if self.table.item(r, 6) else "",
                "available_till": self.table.item(r, 7).text() if self.table.item(r, 7) else "",
                "unavailable_days": self.table.item(r, 8).text() if self.table.item(r, 8) else "",
                "notes": self.table.item(r, 9).text() if self.table.item(r, 9) else ""
            }
            
            # Very basic parse example
            constraints = {}
            if "max 2 consecutive" in data["notes"].lower():
                constraints["max_consecutive"] = 2
                
            data["constraints_json"] = json.dumps(constraints)
            insert_teacher(data)
            
        QMessageBox.information(self, "Saved", "Teacher data saved to database.")
